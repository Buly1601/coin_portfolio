from datetime import date, timedelta
import pandas as pd
import openpyxl as pxl
import colorama

TODAY = date.today()
COIN = 0
MXN = 1

class command:

    def __init__(self, ammount, coin, mxn, conv_ammount=None):
        self.ammount = float(ammount)
        self.coin = coin
        self.file = "portfolio.xlsx"
        self.mxn = float(mxn)
        self.book = pxl.load_workbook(self.file)
        self.date = TODAY.strftime("%d/%m/%Y")
        # TODO -------------------------------
        self.conversion = conv_ammount
        # !TEST ------------------------------
        


    def deposit(self):
        """ 
        Deposit command, receives float number and adds to coin 
        total ammount.
        """
        # read file as pandas dataframe
        file = pd.read_excel(self.file, index_col=[0])

        with pd.ExcelWriter(self.file, engine="openpyxl") as writer:
            # save book as base
            writer.book = self.book
            writer.sheets = dict((ws.title, ws) for ws in self.book.worksheets)

            # if non-existing coin, create column
            if self.coin not in file:
                # add new column
                file[self.coin] = [self.ammount, self.mxn]
                # rename indexes 
                file.rename(index={0:"COIN", 1:"MXN"}, inplace=True)

            else:
                # add ammounts to respective coins
                file[self.coin] += [self.ammount, self.mxn]
            
            # show user action
            print(colorama.Fore.GREEN + f"\n Deposited {str(self.ammount)} to {self.coin}.")
            print(f" ----- \n AMMOUNT OF {self.coin} -> {str(file[self.coin][COIN])} at {self.date} \n ----- ")
            # reset colors
            print(colorama.Style.RESET_ALL)
            # record and save
            file.to_excel(writer, index_label=[0])
            writer.save()

        # record action
        self.record_action("deposit")


    def withdraw(self):
        """ 
        Withdraw command, receives float number and substracts to coin 
        total ammount.
        """
        # read file as pandas dataframe
        file = pd.read_excel(self.file, index_col=[0])

        # check if coin exists
        if self.coin not in file:
            raise Exception(f"No current history of {self.coin}")
        
        # check if currency enough for withdrawal
        if file[self.coin][COIN] < self.ammount:
            raise Exception(f"Not enough currency of {self.coin}, \n ----- \n AVAILABLE -> {file[self.coin][COIN]} \n -----")
        
        else:
            # read excel with openpyxl engine
            with pd.ExcelWriter(self.file, engine="openpyxl") as writer:
                # set file as base
                writer.book = self.book
                writer.sheets = dict((ws.title, ws) for ws in self.book.worksheets)

                # withdraw coin from data
                file[self.coin][COIN] -= self.ammount
                # show user action
                print(colorama.Fore.RED + f"\n Withdrawed {str(self.ammount)} from {self.coin}.")
                print(f" ----- \n AMMOUNT OF {self.coin} -> {str(file[self.coin][COIN])} at {self.date} \n ----- ")
                # reset colors
                print(colorama.Style.RESET_ALL)
                # record and save
                file.to_excel(writer, index_label=[0])
                writer.save()
        
        # record action to coin's history
        self.record_action("withdraw")

    
    def graph(self):
        """
        Graph command, plots a pie graph of the current coins ammounts,
        total balance and balance of each.
        """
        #TODO
        return 

    
    def convert(self):
        """
        Converts ammount of coin to another coin.
        """
        # TODO
        return 
    

    def record_action(self, action):
        """
        Records action, and date in coin's specific sheet.
        """
        # get sheets
        sheets = self.book.sheetnames
        # flag for new coin
        n_flag = False

        # create backbone of dataframe in sheet 
        """
        date(index) | coin_ammount | movement
        """
        backbone = {"coin_ammount":self.ammount, "movement":action, "mxn_ammount":self.mxn}
        b_df = pd.DataFrame(backbone, index=[self.date])

        if self.coin not in sheets:
            # create new sheet 
            self.book.create_sheet(self.coin)
            # save newly created sheet
            self.book.save(self.file)
            # activate flag
            n_flag = True

        with pd.ExcelWriter(self.file, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
            # save base
            # set file as base
            writer.book = self.book
            writer.sheets = dict((ws.title, ws) for ws in self.book.worksheets)

            if n_flag:
                # new listed coin
                b_df.to_excel(writer, sheet_name=self.coin, index=True, index_label=[0])
            else:
                # get coin sheet df
                c_df = pd.read_excel(self.file, sheet_name=self.coin, index_col=[0])
                # append backbone to original df
                new = pd.concat([c_df, b_df])
                new.to_excel(writer, sheet_name=self.coin, index=True, index_label=[0])

            writer.save()

        return


    def total(self, key=None, date_=None):
        """
        Returns total ammount of mxn invested.
        """
        # read file as pandas dataframe
        file = pd.read_excel(self.file)
        
        # set total
        total = 0

        # check for the desired dates
        # only three options, today, week, month
        # start with today
        cell = file.index
        if key == "TODAY":
            for i, col in enumerate(file):
                if i == 0: 
                    continue
                if file[col]["Date"] != TODAY: # TODO DATE CELL
                    break
                total += file[col][MXN]

        elif key == "WEEK":
            day_of_week = TODAY.weekday()
            # if its sunday (6) then we have to check max six days before today
            # up to today
            min_date = abs(0-day_of_week)

            for i, col in enumerate(file):
                if i == 0: 
                    continue
                if file[col]["Date"] > TODAY or file[col]["Date"] < TODAY - timedelta(days=min_date): # TODO DATE CELL
                    break
                total += file[col][MXN]

        elif key == "MONTH":
            curr_month = TODAY.month()
            # if its sunday (6) then we have to check max six days before today
            # up to today

            for i, col in enumerate(file):
                if i == 0: 
                    continue
                if file[col]["Date"] <= TODAY and file[col]["Date"] >= TODAY - timedelta(weeks=4): # TODO DATE CELL
                    break
                total += file[col][MXN]

        else:
            raise Exception(f"Date key '{key}' invalid.")
        
        # show user total mxn
        print(colorama.Fore.BLUE + f"\n ----- \n Total ammount of mxn invested {key.lower()} -> {str(total)}$. \n -----")
        print(colorama.Style.RESET_ALL)